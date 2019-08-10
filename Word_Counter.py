
# coding: utf-8

# In[117]:
import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])

try:
    import pandas as pd
    import numpy as np
    import xlrd

except:
    install('pandas')
    install('numpy')
    install('xlrd')



import pandas as pd
import numpy as np


# In[118]:


excel_df_name = str(input("Please give the name of excel with data.\n"))


# In[119]:


excel_df = pd.read_excel(excel_df_name)
excel_df = excel_df.applymap(str)


# In[120]:


excel_df


# In[121]:


newdf = pd.DataFrame() 
for ColName in excel_df.columns:
    frequencies = excel_df[ColName].apply(lambda x: pd.value_counts(x.split(" "))).sum(axis = 0)
    print(frequencies)
    #newdf = newdf.append(frequencies.to_frame(name = ColName))
    newdf = pd.concat([newdf, frequencies.to_frame(name = ColName)], axis=1)


# In[122]:


newdf


# In[123]:


newdf.to_excel("frequencies_in_"+ excel_df_name)


# In[124]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===savetime===
from datetime import datetime
thetime= datetime.now().strftime("%Y%m%d-%H%M%S")

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

